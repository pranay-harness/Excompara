import argparse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import pandas as pd
import pyfiglet


class ExCompara:
    """
    A class for comparing Excel files and finding fixed CVE_ID values, newly added CVEs, and severity analysis.
    """

    def __init__(self, old_file_path, new_file_path):
        """
        Initialize the ExCompara with the paths to the old and new Excel files.

        Args:
            old_file_path (str): Path to the old Excel file.
            new_file_path (str): Path to the new Excel file.
        """
        self.old_file_path = old_file_path
        self.new_file_path = new_file_path

    def display():
        """
        Method to display tool name, author information and usage
        """
        # Display banner
        banner_text = GREEN_TEXT + pyfiglet.figlet_format("ExCompara") + RESET_TEXT

        # Information formatting
        info = """

        Date: June 2023

        Team: Product Security

        Description: An internal tool that analyzes and provides insights on the changes between two versions of reports.
        
        Note:
        
        - Make sure to run the 'dependencies.py' script to install any missing dependency!

        - Make sure both the files is in same directory.

        """

        # Create the formatted banner text with author information
        banner_text_with_info = f"\033[34m{banner_text}\033[0m\n\n{info}"

        # Add border around the banner text
        banner_border = "+" + "-" * (len(banner_text.split("\n")[0]) + 50) + "+"
        formatted_banner = f"{banner_border}\n{banner_text_with_info}\n{banner_border}"

        return formatted_banner

    def _read_excel_file(self, file_path):
        """
        Read an Excel file and build a mapping of sheet names to CVE_ID values.

        Args:
            file_path (str): Path to the Excel file.

        Returns:
            dict: Mapping of sheet names to CVE_ID values.
        """
        try:
            excel_file = pd.ExcelFile(file_path)
        except FileNotFoundError:
            raise FileNotFoundError(f"File '{file_path}' not found.")

        mapping = {}

        for sheet_name in excel_file.sheet_names:
            if sheet_name != "Sheet1":
                sheet_data = excel_file.parse(sheet_name)

                if "CVE_ID" in sheet_data.columns:
                    cve_ids = sheet_data["CVE_ID"].dropna().tolist()

                    if sheet_name not in mapping:
                        mapping[sheet_name] = []
                    mapping[sheet_name].extend(cve_ids)

        return mapping

    def find_fixed_cves(self):
        """
        Find fixed CVE_ID values for each sheet by comparing the old and new Excel files.

        Returns:
            dict: Fixed CVE_ID values for each sheet.
        """
        old_mapping = self._read_excel_file(self.old_file_path)
        new_mapping = self._read_excel_file(self.new_file_path)

        fixed_cves = {}
        for sheet_name, old_cves in old_mapping.items():
            new_cves = new_mapping.get(sheet_name, [])
            missing_cves = list(set(old_cves) - set(new_cves))
            if missing_cves:
                fixed_cves[sheet_name] = missing_cves

        return fixed_cves

    def find_newly_added_cves(self):
        """
        Find newly added CVE_ID values for each sheet by comparing the old and new Excel files.

        Returns:
            dict: Newly added CVE_ID values for each sheet.
        """
        old_mapping = self._read_excel_file(self.old_file_path)
        new_mapping = self._read_excel_file(self.new_file_path)

        newly_added_cves = {}
        for sheet_name, new_cves in new_mapping.items():
            old_cves = old_mapping.get(sheet_name, [])
            added_cves = list(set(new_cves) - set(old_cves))
            if added_cves:
                newly_added_cves[sheet_name] = added_cves

        return newly_added_cves

    def compare_severity_analysis(self):
        """
        Compare the 'vulnerability_count' sheet in the old and new Excel files to analyze severity differences.

        Returns:
            pd.DataFrame: Difference in severity analysis.
        """
        try:
            old_data = pd.read_excel(
                self.old_file_path, sheet_name="vulnerability_count"
            )
            new_data = pd.read_excel(
                self.new_file_path, sheet_name="vulnerability_count"
            )
        except FileNotFoundError as e:
            raise FileNotFoundError(f"File not found: {e.filename}")

        # Compute the difference by subtracting the new data from the old data
        diff_data = old_data.copy()
        diff_data.iloc[:, 1:] = old_data.iloc[:, 1:] - new_data.iloc[:, 1:]

        return diff_data

    def calculate_cve_changes(self):
        """
        Calculate the percentage decrement in newly introduced CVEs and percentage increment in fixed CVEs.
        Prints the results on the command-line.
        """
        old_mapping = self._read_excel_file(self.old_file_path)
        new_mapping = self._read_excel_file(self.new_file_path)

        distinct_old_cves = set(cve for cves in old_mapping.values() for cve in cves)
        distinct_new_cves = set(cve for cves in new_mapping.values() for cve in cves)
        print(distinct_old_cves)
        print(distinct_new_cves)
        total_old_cves = len(distinct_old_cves)
        total_new_cves = len(distinct_new_cves)
        print(total_old_cves)
        print(total_new_cves)

        percent_increment = (total_new_cves / total_old_cves) * 100 if total_old_cves != 0 else 0

        print(f"Percentage Increment in Newly Introduced CVEs: {percent_increment:.2f}%")


    def generate_analysis_report(self):
        """
        Generate an analysis report in Excel format with fixed and newly added CVEs, and severity analysis.

        Returns:
            None
        """
        # Find fixed CVEs
        fixed_cves = self.find_fixed_cves()

        # Find newly added CVEs
        newly_added_cves = self.find_newly_added_cves()

        # Compare severity analysis
        severity_diff = self.compare_severity_analysis()

        # Create a workbook and select the active sheet
        workbook = Workbook()
        sheet = workbook.active

        # Set up styles for formatting
        box_fill = PatternFill(
            start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
        )
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        cell_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        header_font = Font(name="Calibri", bold=True, size=14)

        # Write fixed CVEs
        sheet["A1"] = "Fixed CVEs"
        sheet["A1"].font = header_font
        sheet["A1"].fill = box_fill
        sheet["A1"].border = border
        sheet["A1"].alignment = cell_alignment

        sheet["A2"] = "CVE Fixed"
        sheet["A2"].font = header_font
        sheet["A2"].fill = box_fill
        sheet["A2"].border = border
        sheet["A2"].alignment = cell_alignment

        sheet["B2"] = "Image"
        sheet["B2"].font = header_font
        sheet["B2"].fill = box_fill
        sheet["B2"].border = border
        sheet["B2"].alignment = cell_alignment

        row = 3
        for sheet_name, cves in fixed_cves.items():
            for cve in cves:
                if pd.notnull(cve):
                    sheet.cell(row=row, column=1).value = cve
                    sheet.cell(row=row, column=2).value = sheet_name
                    sheet.cell(row=row, column=1).border = border
                    sheet.cell(row=row, column=2).border = border
                    row += 1

        # Add space between the two boxes
        row += 10

        # Write newly added CVEs
        sheet["A" + str(row)] = "Newly Added CVEs"
        sheet["A" + str(row)].font = header_font
        sheet["A" + str(row)].fill = box_fill
        sheet["A" + str(row)].border = border
        sheet["A" + str(row)].alignment = cell_alignment

        sheet["A" + str(row + 1)] = "CVE Added"
        sheet["A" + str(row + 1)].font = header_font
        sheet["A" + str(row + 1)].fill = box_fill
        sheet["A" + str(row + 1)].border = border
        sheet["A" + str(row + 1)].alignment = cell_alignment

        sheet["B" + str(row + 1)] = "Image"
        sheet["B" + str(row + 1)].font = header_font
        sheet["B" + str(row + 1)].fill = box_fill
        sheet["B" + str(row + 1)].border = border
        sheet["B" + str(row + 1)].alignment = cell_alignment

        row += 2
        for sheet_name, cves in newly_added_cves.items():
            for cve in cves:
                if pd.notnull(cve):
                    sheet.cell(row=row, column=1).value = cve
                    sheet.cell(row=row, column=2).value = sheet_name
                    sheet.cell(row=row, column=1).border = border
                    sheet.cell(row=row, column=2).border = border
                    row += 1

        # Add space between the second and third box
        row += 10

        # Write severity analysis
        sheet["A" + str(row)] = "Severity Analysis"
        sheet["A" + str(row)].font = header_font
        sheet["A" + str(row)].fill = box_fill
        sheet["A" + str(row)].border = border
        sheet["A" + str(row)].alignment = cell_alignment

        # Write severity analysis headers
        severity_headers = list(severity_diff.columns)
        for i, header in enumerate(severity_headers):
            col = chr(ord("A") + i)
            sheet[col + str(row + 1)] = header
            sheet[col + str(row + 1)].font = header_font
            sheet[col + str(row + 1)].fill = box_fill
            sheet[col + str(row + 1)].border = border
            sheet[col + str(row + 1)].alignment = cell_alignment

        # Write severity analysis data
        for i, (_, row_data) in enumerate(severity_diff.iterrows()):
            for j, value in enumerate(row_data):
                col = chr(ord("A") + j)
                sheet[col + str(row + 2 + i)] = value
                sheet[col + str(row + 2 + i)].border = border

        # Adjust column widths
        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 20

        # Save the workbook
        try:
            workbook.save(filename="analysis_report.xlsx")
            print(
                "-----------------------------------------------------------------------------------------------------------"
            )
            print(
                GREEN_TEXT
                + "\n                                    Analysis report generated successfully.\n             "
                + RESET_TEXT
            )
            print(
                "-----------------------------------------------------------------------------------------------------------"
            )
        except PermissionError:
            print(
                RED_TEXT
                + "Error: Permission denied. Please close the analysis_report.xlsx file and try again."
                + RESET_TEXT
            )
        except Exception as e:
            print("Error:", str(e))


if __name__ == "__main__":
    # ANSI codes
    RED_TEXT = "\033[31m"
    GREEN_TEXT = "\033[32m"
    RESET_TEXT = "\033[0m"

    print(ExCompara.display())

    # Parse command line arguments
    parser = argparse.ArgumentParser(description="Excel Comparator")
    parser.add_argument("old_file", type=str, help="Path to the old Excel file")
    parser.add_argument("new_file", type=str, help="Path to the new Excel file")
    args = parser.parse_args()

    # Create an instance of ExCompara
    comparator = ExCompara(args.old_file, args.new_file)

    comparator.calculate_cve_changes()

    # Generate the analysis report
    comparator.generate_analysis_report()